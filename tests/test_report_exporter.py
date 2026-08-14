"""报告导出模块单元测试。"""
from __future__ import annotations

from datetime import date

import pandas as pd

from agent_platform.finance.analysis import (
    DISCLAIMER,
    SecurityAnalysisResult,
)
from agent_platform.finance.report_exporter import to_excel_bytes, to_html_bytes


def _make_result() -> SecurityAnalysisResult:
    """构造一个最小有效的 SecurityAnalysisResult 供测试用。"""
    df = pd.DataFrame({
        "market": ["上交所"] * 5,
        "symbol": ["600519"] * 5,
        "name": ["贵州茅台"] * 5,
        "date": [date(2025, 1, i + 2) for i in range(5)],
        "open": [100.0, 101.0, 102.0, 99.0, 103.0],
        "high":  [102.0, 103.0, 104.0, 101.0, 105.0],
        "low":   [99.0, 100.0, 101.0, 98.0, 102.0],
        "close": [101.0, 102.0, 100.0, 103.0, 104.0],
        "volume": [10000] * 5,
        "source": ["test"] * 5,
        "updated_at": ["2025-01-06"] * 5,
        "ma5":  [101.0, 101.5, 101.0, 101.5, 102.0],
        "ma20": [101.0] * 5,
        "macd": [0.1, 0.2, 0.3, 0.4, 0.5],
        "macd_signal": [0.05, 0.1, 0.15, 0.2, 0.25],
        "macd_hist": [0.1, 0.2, 0.3, 0.4, 0.5],
        "rsi": [50.0, 55.0, 45.0, 60.0, 55.0],
        "bb_upper": [105.0] * 5,
        "bb_middle": [101.0] * 5,
        "bb_lower": [97.0] * 5,
        "kdj_k": [50.0, 55.0, 60.0, 65.0, 70.0],
        "kdj_d": [48.0, 52.0, 57.0, 62.0, 67.0],
        "kdj_j": [54.0, 61.0, 66.0, 71.0, 76.0],
        "atr": [1.5, 1.6, 1.4, 1.7, 1.5],
        "cci": [50.0, 60.0, 40.0, 80.0, 70.0],
        "volume_ma5": [10000.0] * 5,
        "ema12": [101.0, 101.5, 101.0, 101.5, 102.0],
        "ema26": [100.5, 101.0, 100.5, 101.0, 101.5],
    })
    return SecurityAnalysisResult(
        market="上交所",
        symbol="600519",
        name="贵州茅台",
        start_date="2025-01-02",
        end_date="2025-01-06",
        source="test",
        updated_at="2025-01-06",
        total_return_pct=4.0,
        annualized_volatility_pct=15.0,
        max_drawdown_pct=-2.0,
        latest_close=104.0,
        latest_ma5=102.0,
        latest_ma20=101.0,
        latest_rsi=55.0,
        latest_macd=0.5,
        latest_macd_signal=0.25,
        latest_bb_upper=105.0,
        latest_bb_lower=97.0,
        latest_bb_position_pct=87.5,
        latest_kdj_k=70.0,
        latest_kdj_d=67.0,
        latest_kdj_j=76.0,
        latest_atr=1.5,
        latest_cci=70.0,
        latest_ema12=102.0,
        latest_ema26=101.5,
        disclaimer=DISCLAIMER,
        price_history=df,
        data_status="offline_sample",
        fallback_reason=None,
    )


class TestExcelExport:
    def test_returns_bytes(self) -> None:
        data = to_excel_bytes(_make_result())
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_two_sheets(self) -> None:
        import io
        from openpyxl import load_workbook

        data = to_excel_bytes(_make_result())
        wb = load_workbook(io.BytesIO(data))
        assert "分析摘要" in wb.sheetnames
        assert "历史行情" in wb.sheetnames

    def test_summary_contains_symbol(self) -> None:
        import io
        from openpyxl import load_workbook

        data = to_excel_bytes(_make_result())
        wb = load_workbook(io.BytesIO(data))
        ws = wb["分析摘要"]
        # 第一列第二行应该是证券代码值
        found = any(
            str(cell.value) == "600519"
            for row in ws.iter_rows()
            for cell in row
        )
        assert found, "分析摘要 sheet 中应包含证券代码 600519"


class TestHtmlExport:
    def test_returns_bytes(self) -> None:
        data = to_html_bytes(_make_result())
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_contains_disclaimer(self) -> None:
        data = to_html_bytes(_make_result())
        html = data.decode("utf-8")
        assert "不构成投资建议" in html

    def test_contains_chart_elements(self) -> None:
        data = to_html_bytes(_make_result())
        html = data.decode("utf-8")
        # Plotly 图表输出要么包含 <script> 要么包含 <div class="plotly
        assert ("<script>" in html) or ("plotly" in html.lower())

    def test_chart_titles_are_separate_from_plot_legends(self) -> None:
        html = to_html_bytes(_make_result()).decode("utf-8")
        assert "<h3>K线、均线与布林带</h3>" in html
        assert "<h3>MACD（12/26/9）</h3>" in html
        assert "<h3>KDJ（9/3/3）</h3>" in html
        assert "<h3>RSI（14）</h3>" in html
        assert "margin\":{\"l\":52,\"r\":84" in html

    def test_html_export_embeds_plotly_for_offline_viewing(self) -> None:
        html = to_html_bytes(_make_result()).decode("utf-8")
        assert '<script src="https://cdn.plot.ly' not in html
        assert "plotly.js v" in html
